#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON to Excel 변환기
new_data.json 파일을 읽어서 Excel 리포트 생성
"""

import json
import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# UTF-8 인코딩 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

def load_new_data(file_path="output/new_data.json"):
    """new_data.json 파일 로드 (AI 요약이 포함된 신규 데이터)"""
    try:
        if not os.path.exists(file_path):
            print(f"❌ {file_path} 파일이 없습니다.")
            return None
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print(f"📂 {len(data)}개 공고 로드 완료")
        return data
        
    except Exception as e:
        print(f"❌ 파일 로드 실패: {str(e)}")
        return None

def create_excel_report(announcements, output_file=None):
    """Excel 리포트 생성"""
    try:
        print("\n📊 Excel 리포트 생성 시작...")
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "신규공고"
        
        # 스타일 정의
        header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='맑은 고딕', size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 항목명 (세로로 배치)
        field_names = [
            "공고명", "부처명", "현황", "접수일", "마감일",
            "사업목적", "지원내용", "지원규모", "신청대상", "주요특징", "전체요약", "처리상태", "상세URL"
        ]
        
        # 첫 번째 열에 항목명 입력
        ws.cell(row=1, column=1, value="항목").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=1).border = border
        
        for row, field_name in enumerate(field_names, 2):
            cell = ws.cell(row=row, column=1, value=field_name)
            cell.font = Font(name='맑은 고딕', size=10, bold=True)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 첫 번째 열 너비
        ws.column_dimensions['A'].width = 15
        
        # 각 공고를 열로 배치
        for col_idx, announcement in enumerate(announcements, 2):  # 2열부터 시작
            ai_summary = announcement.get('ai_요약', {})
            
            # 헤더: 공고 번호
            header_cell = ws.cell(row=1, column=col_idx, value=f"공고 {col_idx - 1}")
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = border
            
            # 데이터
            values = [
                announcement.get('공고명', ''),
                announcement.get('부처명', ''),
                announcement.get('현황', ''),
                announcement.get('접수일', ''),
                announcement.get('마감일', ''),
                ai_summary.get('사업목적', ''),
                ai_summary.get('지원내용', ''),
                ai_summary.get('지원규모', ''),
                ai_summary.get('신청대상', ''),
                ai_summary.get('주요특징', ''),
                ai_summary.get('전체요약', ''),
                announcement.get('처리상태', ''),
                announcement.get('상세_URL', '')
            ]
            
            for row_idx, value in enumerate(values, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # 특수 처리
                if row_idx == 4:  # 현황
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if '접수중' in str(value):
                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    elif '마감' in str(value):
                        cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                elif row_idx == 12:  # 전체요약 - 행 높이 증가
                    ws.row_dimensions[row_idx].height = 80
                elif row_idx == 13:  # 처리상태
                    if value:  # 예외 상황이 있는 경우
                        cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                        cell.font = Font(name='맑은 고딕', size=10, bold=True, color='856404')
                elif row_idx == 14:  # URL
                    if value:
                        cell.hyperlink = value
                        cell.font = Font(name='맑은 고딕', size=10, color='0563C1', underline='single')
                        cell.value = "바로가기"
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 조정
            ws.column_dimensions[get_column_letter(col_idx)].width = 50
        
        # 행 높이 조정
        for row in range(1, len(field_names) + 2):
            ws.row_dimensions[row].height = 40
        
        # 파일명 생성
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"NTIS_신규공고_{timestamp}.xlsx"
        
        # 저장 경로 설정
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        filepath = os.path.join(output_dir, output_file)
        
        # 파일 저장
        wb.save(filepath)
        
        print(f"✅ Excel 리포트 생성 완료!")
        print(f"   파일: {filepath}")
        print(f"   공고 수: {len(announcements)}건")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Excel 리포트 생성 실패: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """메인 함수"""
    print("=" * 60)
    print("JSON to Excel 변환기")
    print("=" * 60)
    
    # 1. new_data.json 로드
    announcements = load_new_data()
    if not announcements:
        print("처리할 데이터가 없습니다.")
        return
    
    # 2. Excel 리포트 생성
    filepath = create_excel_report(announcements)
    
    if filepath:
        print(f"\n🎉 변환 완료!")
        print(f"Excel 파일을 확인하세요: {filepath}")
    else:
        print("\n❌ 변환 실패!")

if __name__ == "__main__":
    main()

