#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 리포트 생성기
NTIS 공고 데이터를 Excel 파일로 생성
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os

class ExcelReportGenerator:
    """Excel 리포트 생성 클래스"""
    
    def __init__(self):
        """초기화"""
        self.workbook = None
        self.worksheet = None
        
        # 스타일 정의
        self.header_font = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.data_font = Font(name='맑은 고딕', size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        print("Excel 리포트 생성기 초기화 완료")
    
    def create_workbook(self, title="NTIS 공고 리포트"):
        """새 워크북 생성"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "공고목록"
        
        print(f"워크북 생성: {title}")
        return self.workbook
    
    def setup_headers(self):
        """헤더 설정"""
        headers = [
            "순번", "현황", "공고명", "부처명", "접수일", "마감일", "남은일수", "링크"
        ]
        
        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 컬럼 너비 조정
        column_widths = {
            'A': 8,   # 순번
            'B': 12,  # 현황
            'C': 50,  # 공고명
            'D': 25,  # 부처명
            'E': 15,  # 접수일
            'F': 15,  # 마감일
            'G': 12,  # 남은일수
            'H': 80   # 링크
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
        
        print("헤더 설정 완료")
    
    def add_announcement(self, row_num, announcement):
        """공고 데이터 추가"""
        try:
            # 데이터 추출
            순번 = announcement.get('순번', row_num - 1)
            현황 = announcement.get('현황', '')
            공고명 = announcement.get('공고명', '')
            부처명 = announcement.get('부처명', '')
            접수일 = announcement.get('접수일', '')
            마감일 = announcement.get('마감일', '')
            남은일수 = announcement.get('남은일수_계산', announcement.get('남은일수', ''))
            링크 = announcement.get('링크', '')
            
            # 데이터 입력
            data = [순번, 현황, 공고명, 부처명, 접수일, 마감일, 남은일수, 링크]
            
            for col, value in enumerate(data, 1):
                cell = self.worksheet.cell(row=row_num, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.border
                
                # 정렬 설정
                if col == 1 or col == 7:  # 순번, 남은일수
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 2:  # 현황
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 현황에 따른 색상 설정
                    if '접수중' in str(value):
                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    elif '마감' in str(value):
                        cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                elif col == 8:  # 링크
                    if value:
                        cell.hyperlink = value
                        cell.font = Font(name='맑은 고딕', size=10, color='0563C1', underline='single')
                        cell.value = "링크 바로가기"
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 남은일수에 따른 행 색상 설정
            if isinstance(남은일수, int):
                if 남은일수 <= 3:  # 3일 이하
                    for col in range(1, 9):
                        if col != 2:  # 현황 컬럼은 제외
                            self.worksheet.cell(row=row_num, column=col).fill = PatternFill(
                                start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'
                            )
                elif 남은일수 <= 7:  # 7일 이하
                    for col in range(1, 9):
                        if col != 2:  # 현황 컬럼은 제외
                            self.worksheet.cell(row=row_num, column=col).fill = PatternFill(
                                start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
                            )
            
        except Exception as e:
            print(f"공고 데이터 추가 실패 (행 {row_num}): {str(e)}")
    
    def add_summary_sheet(self, summary_data):
        """요약 시트 추가"""
        try:
            # 새 시트 생성
            summary_sheet = self.workbook.create_sheet(title="요약")
            
            # 요약 데이터 작성
            summary_sheet['A1'] = "NTIS 공고 크롤링 요약"
            summary_sheet['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
            summary_sheet.merge_cells('A1:D1')
            
            # 요약 정보
            summary_info = [
                ["생성일시", datetime.now().strftime("%Y년 %m월 %d일 %H:%M:%S")],
                ["검색 키워드", summary_data.get('keyword', 'AI')],
                ["총 추출 공고", f"{summary_data.get('extracted_count', 0)}건"],
                ["신청 가능 공고", f"{summary_data.get('filtered_count', 0)}건"],
                ["신규 공고", f"{summary_data.get('new_count', 0)}건"],
                ["중복 공고", f"{summary_data.get('duplicate_count', 0)}건"]
            ]
            
            for row, (label, value) in enumerate(summary_info, 3):
                summary_sheet.cell(row=row, column=1, value=label).font = Font(name='맑은 고딕', size=12, bold=True)
                summary_sheet.cell(row=row, column=2, value=value).font = Font(name='맑은 고딕', size=12)
            
            # 컬럼 너비 조정
            summary_sheet.column_dimensions['A'].width = 20
            summary_sheet.column_dimensions['B'].width = 30
            
            print("요약 시트 추가 완료")
            
        except Exception as e:
            print(f"요약 시트 추가 실패: {str(e)}")
    
    def generate_report(self, announcements, summary_data=None, filename=None):
        """전체 리포트 생성"""
        try:
            # 워크북 생성
            self.create_workbook()
            
            # 헤더 설정
            self.setup_headers()
            
            # 공고 데이터 추가
            for i, announcement in enumerate(announcements, 2):  # 2행부터 시작 (1행은 헤더)
                self.add_announcement(i, announcement)
            
            # 요약 시트 추가
            if summary_data:
                self.add_summary_sheet(summary_data)
            
            # 파일명 생성
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"NTIS_공고리포트_{timestamp}.xlsx"
            
            # 저장 경로 설정
            output_dir = "output"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            filepath = os.path.join(output_dir, filename)
            
            # 파일 저장
            self.workbook.save(filepath)
            
            print(f"Excel 리포트 생성 완료: {filepath}")
            print(f"총 {len(announcements)}건의 공고가 포함되었습니다.")
            
            return filepath
            
        except Exception as e:
            print(f"Excel 리포트 생성 실패: {str(e)}")
            return None
    
    def generate_new_announcements_report(self, new_announcements, all_data, filename=None):
        """신규 공고만 포함된 리포트 생성"""
        try:
            if not new_announcements:
                print("신규 공고가 없어 Excel 리포트를 생성하지 않습니다.")
                return None
            
            # 요약 데이터 준비
            summary_data = {
                'keyword': all_data.get('keyword', 'AI'),
                'extracted_count': all_data.get('extracted_count', 0),
                'filtered_count': all_data.get('filtered_count', 0),
                'new_count': len(new_announcements),
                'duplicate_count': all_data.get('duplicate_count', 0)
            }
            
            # 파일명에 신규 표시
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"NTIS_신규공고_{timestamp}.xlsx"
            
            return self.generate_report(new_announcements, summary_data, filename)
            
        except Exception as e:
            print(f"신규 공고 Excel 리포트 생성 실패: {str(e)}")
            return None

def main():
    """테스트용 메인 함수"""
    print("Excel 리포트 생성기 테스트")
    
    # 테스트 데이터
    test_announcements = [
        {
            "순번": 1,
            "현황": "접수중",
            "공고명": "2024년 AI 기술개발 지원사업",
            "부처명": "과학기술정보통신부",
            "접수일": "2024.01.01",
            "마감일": "2024.12.31",
            "남은일수_계산": 5,
            "링크": "https://www.ntis.go.kr/example1"
        },
        {
            "순번": 2,
            "현황": "접수중",
            "공고명": "스마트시티 솔루션 개발 과제",
            "부처명": "국토교통부",
            "접수일": "2024.02.01",
            "마감일": "2024.12.15",
            "남은일수_계산": 15,
            "링크": "https://www.ntis.go.kr/example2"
        }
    ]
    
    # Excel 생성기 테스트
    generator = ExcelReportGenerator()
    
    summary_data = {
        'keyword': 'AI',
        'extracted_count': 50,
        'filtered_count': 25,
        'new_count': 2,
        'duplicate_count': 23
    }
    
    filepath = generator.generate_report(test_announcements, summary_data, "테스트_리포트.xlsx")
    
    if filepath:
        print(f"테스트 성공! 파일 생성: {filepath}")
    else:
        print("테스트 실패!")

if __name__ == "__main__":
    main()
